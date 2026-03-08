import sys
import openpyxl
from openpyxl.styles import Font
import os

def write_to_excel(file_path, sheet_name, cell_ref, value):
    try:
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()
            
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet" and not os.path.exists(file_path):
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

        final_value = value
        if not str(value).startswith('='):
            try:
                if '.' in str(value): final_value = float(value)
                else: final_value = int(value)
            except ValueError:
                pass

        # Write value
        target_cell = ws[cell_ref]
        target_cell.value = final_value
        
        # --- [CRITICAL] Apply Thai Sarabun Font ---
        thai_font = Font(name='TH SarabunPSK', size=16)
        target_cell.font = thai_font
        # ------------------------------------------
        
        wb.save(file_path)
        print(f"✅ Successfully wrote '{value}' to [{sheet_name}]{cell_ref} with TH SarabunPSK 16pt")
        
    except Exception as e:
        print(f"❌ Error writing to Excel: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 5:
        print("Usage: python write_excel.py <file_path> <sheet_name> <cell_ref> <value_or_formula>")
        sys.exit(1)
    write_to_excel(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
